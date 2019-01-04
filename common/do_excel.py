
# 读取测试数据
import openpyxl
import os
from api_auto_test.common import project_path
from api_auto_test.common.config import Conf
class Case:  # 测试用例封装类
    def __init__(self):
        self.case_id = None
        self.title = None
        self.http_method = None
        self.url = None
        self.params = None
        self.expected = None
        self.actual = None
        self.result = None

class DoExcel:  # excel操作封装类
    def __init__(self, filename):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(filename=filename)

    def get_case(self,sheet_name):
        option = Conf().get_option_str('API', 'url_pre')  # 获取要执行的配置文件url数据
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        case_data = []
        for i in range(2, max_row+1):
            case = Case()  # 实例化一个case对象，用例存放测试数据
            case.case_id = sheet.cell(i,1).value
            case.title = sheet.cell(i, 2).value
            case.http_method = sheet.cell(i, 3).value
            case.url = os.path.join(option,sheet.cell(i, 4).value)  # 利用os模块的拼接方式，拼接url路径
            # print('case.url=',case.url)
            case.params = sheet.cell(i, 5).value
            case.expected = sheet.cell(i, 6).value
            case_data.append(case)
        return case_data  # 这里的返回值是一个对象列表

    # 获取所有表单
    def sheet_namme(self):   # 这里获取到的sheet表单给get_case方法用
        return self.workbook.sheetnames

    def write_return(self,sheet_name,case_id,actual,result):
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        for i in range(2, max_row+1):
            if sheet.cell(i, 1).value == case_id:
                sheet.cell(i, 7).value = actual
                sheet.cell(i, 8).value = result
                self.workbook.save(self.filename)
                break

if __name__ == '__main__':
    # option = Conf().get_option_str('API', 'url_pre')  # 获取要执行的配置文件url数据
    # 从excel里面获取所有的表单
    do_excel = DoExcel(project_path.testCase_path)
    sheetnames = do_excel.sheet_namme()
    print(sheetnames)
    for j in sheetnames:
        case_data = do_excel.get_case(j)
        print(j)
        for i in case_data:
            print(i.__dict__)#打印每个对象里的测试数据


